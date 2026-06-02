import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime
import os
import sys
try:
    from PIL import Image, ImageTk
    _PIL = True
except ImportError:
    _PIL = False


def get_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def get_resource(filename: str) -> str:
    """Localiza arquivo bundled pelo PyInstaller (_MEIPASS) ou na pasta do script."""
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
        return os.path.join(meipass, filename)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)


EXCEL_FILE = os.path.join(get_base_dir(), "controle_horas_extras.xlsx")

# Formato das colunas Excel (7 colunas):
# [0] Data Início  [1] Hora Início  [2] Data Fim  [3] Hora Fim
# [4] Duração hh:mm:ss  [5] Duração h decimal  [6] Observações

C_BG      = "#F4F6F9"
C_HEADER  = "#1A252F"
C_PRIMARY = "#2C3E50"
C_GREEN   = "#1E8449"
C_RED     = "#C0392B"
C_BLUE    = "#2E86C1"
C_GRAY    = "#BDC3C7"
C_TEXT    = "#2C3E50"
C_MUTED   = "#7F8C8D"
C_WHITE   = "#FFFFFF"
C_ROW_ALT = "#EAF2FB"

_THEMES = {
    "light": dict(
        bg="#F4F6F9", header="#1A252F", primary="#2C3E50",
        green="#1E8449", red="#C0392B", blue="#2E86C1",
        gray="#BDC3C7", text="#2C3E50", muted="#7F8C8D",
        white="#FFFFFF", row_alt="#EAF2FB", obs_bg="#FDFEFE",
        tab_inactive="#D5D8DC", tab_fg="#2C3E50",
        border="#D5D8DC",
    ),
    "dark": dict(
        bg="#1E1E2E", header="#11111B", primary="#CDD6F4",
        green="#1E8449", red="#C0392B", blue="#89B4FA",
        gray="#45475A", text="#CDD6F4", muted="#9399B2",
        white="#252537", row_alt="#2A2A3E", obs_bg="#252537",
        tab_inactive="#313244", tab_fg="#CDD6F4",
        border="#45475A",
    ),
}


class ControleHE:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Controle de HE by Bena")
        self.root.geometry("780x700")
        self.root.resizable(False, False)
        self.root.configure(bg=C_BG)

        self.inicio: datetime.datetime | None = None
        self.timer_running = False
        self.elapsed_seconds = 0
        self._PLACEHOLDER = "Descreva a atividade realizada durante este período..."
        self._is_dark = False

        self._setup_icon()
        self._apply_styles()
        self._build_ctx_menu()
        self._build_header()
        self._build_notebook()
        self._migrate_excel_if_needed()
        self._load_history()

    # ── Ícone da janela ──────────────────────────────────────────
    def _load_logo_src(self) -> "Image.Image | None":
        """Carrega a imagem fonte de alta resolução (PNG preferido, ICO como fallback)."""
        if not _PIL:
            return None
        for name in ("Icon_maior.png", "icon.ico"):
            path = get_resource(name)
            if os.path.exists(path):
                try:
                    img = Image.open(path).convert("RGBA")
                    # Centraliza em quadrado (recorta sombra/margens)
                    w, h = img.size
                    side = min(w, h)
                    img = img.crop(((w-side)//2, (h-side)//2,
                                    (w+side)//2, (h+side)//2))
                    return img
                except Exception:
                    continue
        return None

    def _setup_icon(self):
        src = self._load_logo_src()
        if src is not None:
            try:
                self._icon_64 = ImageTk.PhotoImage(src.resize((64, 64), Image.LANCZOS))
                self._icon_32 = ImageTk.PhotoImage(src.resize((32, 32), Image.LANCZOS))
                self.root.wm_iconphoto(True, self._icon_64, self._icon_32)
                return
            except Exception:
                pass
        try:
            self.root.iconbitmap(get_resource("icon.ico"))
        except Exception:
            pass

    # ── Estilos ttk ──────────────────────────────────────────────
    def _apply_styles(self):
        t = _THEMES["dark" if self._is_dark else "light"]
        s = ttk.Style()
        s.theme_use("clam")
        s.configure("TNotebook", background=t["bg"], borderwidth=0)
        s.configure("TNotebook.Tab", font=("Segoe UI", 10, "bold"),
                    padding=[18, 8], background=t["tab_inactive"],
                    foreground=t["tab_fg"])
        s.map("TNotebook.Tab",
              background=[("selected", t["primary"] if not self._is_dark else "#313244")],
              foreground=[("selected", t["white"] if not self._is_dark else t["primary"])])
        for name, hc in (("HE.Treeview", t["primary"] if not self._is_dark else "#313244"),
                         ("Sum.Treeview", t["blue"])):
            s.configure(name, background=t["white"], fieldbackground=t["white"],
                        foreground=t["text"], font=("Segoe UI", 9), rowheight=24)
            s.configure(f"{name}.Heading",
                        background=t["header"] if self._is_dark else (C_PRIMARY if name == "HE.Treeview" else C_BLUE),
                        foreground="#CDD6F4" if self._is_dark else "#FFFFFF",
                        font=("Segoe UI", 9, "bold"), relief="flat")
            s.map(f"{name}.Heading", background=[("active", t["blue"])])

    # ── Menu de contexto ─────────────────────────────────────────
    def _build_ctx_menu(self):
        self.ctx_menu = tk.Menu(self.root, tearoff=0,
                                font=("Segoe UI", 10),
                                bg=C_WHITE, fg=C_TEXT,
                                activebackground=C_BLUE,
                                activeforeground=C_WHITE,
                                bd=1, relief=tk.SOLID)
        self.ctx_menu.add_command(label="  Editar registro",
                                  command=self._editar_registro)
        self.ctx_menu.add_separator()
        self.ctx_menu.add_command(label="  Excluir registro",
                                  command=self.excluir_registro)

    def _on_right_click(self, event):
        row = self.tree.identify_row(event.y)
        if row:
            self.tree.selection_set(row)
            try:
                self.ctx_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.ctx_menu.grab_release()

    # ── Cabeçalho ────────────────────────────────────────────────
    def _build_header(self):
        hdr = tk.Frame(self.root, bg=C_HEADER, pady=10)
        hdr.pack(fill=tk.X)

        # Logo BENA à esquerda — usa PNG de alta resolução
        src = self._load_logo_src()
        if src is not None:
            try:
                self._hdr_logo = ImageTk.PhotoImage(
                    src.resize((52, 52), Image.LANCZOS))
                tk.Label(hdr, image=self._hdr_logo, bg=C_HEADER,
                         padx=14).pack(side=tk.LEFT)
            except Exception:
                pass

        # Botão dark mode à direita
        self.btn_theme = tk.Button(
            hdr, text="🌙", font=("Segoe UI", 14),
            bg=C_HEADER, fg="#AAB7B8",
            activebackground=C_HEADER, activeforeground="#FFFFFF",
            relief=tk.FLAT, cursor="hand2", bd=0,
            command=self._toggle_theme)
        self.btn_theme.pack(side=tk.RIGHT, padx=(0, 16))

        # Títulos centralizados
        mid = tk.Frame(hdr, bg=C_HEADER)
        mid.pack(expand=True)
        tk.Label(mid, text="CONTROLE DE HORAS EXTRAS",
                 font=("Segoe UI", 15, "bold"), fg=C_WHITE, bg=C_HEADER).pack()
        tk.Label(mid, text="Registre, acompanhe e exporte suas horas",
                 font=("Segoe UI", 9), fg="#AAB7B8", bg=C_HEADER).pack()

    # ── Dark / Light mode ────────────────────────────────────────
    def _toggle_theme(self):
        t_old = _THEMES["dark" if self._is_dark else "light"]
        self._is_dark = not self._is_dark
        t_new = _THEMES["dark" if self._is_dark else "light"]

        # Mapa de cores: hex_antigo → hex_novo (lower-case)
        cmap: dict[str, str] = {}
        for k in t_new:
            old_v = t_old[k].lower()
            new_v = t_new[k]
            cmap[old_v] = new_v

        # Atualiza globais (usados nas criações dinâmicas de janelas)
        global C_BG, C_HEADER, C_PRIMARY, C_BLUE, C_GRAY
        global C_TEXT, C_MUTED, C_WHITE, C_ROW_ALT
        C_BG      = t_new["bg"]
        C_HEADER  = t_new["header"]
        C_PRIMARY = t_new["primary"]
        C_BLUE    = t_new["blue"]
        C_GRAY    = t_new["gray"]
        C_TEXT    = t_new["text"]
        C_MUTED   = t_new["muted"]
        C_WHITE   = t_new["white"]
        C_ROW_ALT = t_new["row_alt"]

        # Recolore todos os widgets
        self._retheme(self.root, cmap)

        # Estilos ttk
        self._apply_styles()

        # Cores específicas pós-recolor
        self.tree.tag_configure("alt", background=t_new["row_alt"])
        if hasattr(self, "tree_res"):
            self.tree_res.tag_configure("alt", background=t_new["row_alt"])
        self.ctx_menu.configure(
            bg=t_new["white"], fg=t_new["text"],
            activebackground=t_new["blue"])

        # Atualiza ícone do botão
        self.btn_theme.config(
            text="☀" if self._is_dark else "🌙",
            bg=t_new["header"], activebackground=t_new["header"])

    def _retheme(self, widget, cmap: dict[str, str]):
        """Percorre a árvore de widgets atualizando cores."""
        for child in widget.winfo_children():
            wtype = type(child).__name__
            for attr in ("background", "bg"):
                try:
                    c = str(child.cget(attr)).lower()
                    if c in cmap:
                        child.configure(**{attr: cmap[c]})
                    break
                except Exception:
                    pass
            if wtype not in ("Scrollbar",):
                for attr in ("foreground", "fg"):
                    try:
                        c = str(child.cget(attr)).lower()
                        if c in cmap:
                            child.configure(**{attr: cmap[c]})
                        break
                    except Exception:
                        pass
            for attr in ("highlightbackground", "insertbackground"):
                try:
                    c = str(child.cget(attr)).lower()
                    if c in cmap:
                        child.configure(**{attr: cmap[c]})
                except Exception:
                    pass
            self._retheme(child, cmap)

    # ── Notebook ─────────────────────────────────────────────────
    def _build_notebook(self):
        wrap = tk.Frame(self.root, bg=C_BG, padx=20, pady=12)
        wrap.pack(fill=tk.BOTH, expand=True)
        self.nb = ttk.Notebook(wrap)
        self.nb.pack(fill=tk.BOTH, expand=True)
        tab1 = tk.Frame(self.nb, bg=C_BG)
        tab2 = tk.Frame(self.nb, bg=C_BG)
        self.nb.add(tab1, text="  Registrar  ")
        self.nb.add(tab2, text="  Resumo  ")
        self.nb.bind("<<NotebookTabChanged>>", self._on_tab_change)
        self._build_tab_registro(tab1)
        self._build_tab_resumo(tab2)

    # ══════════════════════════════════════════════════════════════
    # ABA 1 — REGISTRAR
    # ══════════════════════════════════════════════════════════════
    def _build_tab_registro(self, parent):
        f = tk.Frame(parent, bg=C_BG, pady=10)
        f.pack(fill=tk.BOTH, expand=True)
        self._build_timer_card(f)
        self._build_buttons(f)
        self._build_obs(f)
        self._build_history(f)

    def _build_timer_card(self, parent):
        card = tk.Frame(parent, bg=C_WHITE,
                        highlightbackground="#D5D8DC", highlightthickness=1)
        card.pack(fill=tk.X, pady=(0, 10))
        left = tk.Frame(card, bg=C_WHITE, padx=18, pady=14)
        left.pack(side=tk.LEFT, fill=tk.Y)
        self.status_dot = tk.Label(left, text="●", font=("Segoe UI", 14),
                                   fg=C_GRAY, bg=C_WHITE)
        self.status_dot.pack(anchor=tk.W)
        self.status_lbl = tk.Label(left, text="Aguardando início",
                                   font=("Segoe UI", 10), fg=C_MUTED, bg=C_WHITE)
        self.status_lbl.pack(anchor=tk.W)
        right = tk.Frame(card, bg=C_WHITE, padx=18, pady=14)
        right.pack(side=tk.RIGHT, fill=tk.Y)
        tk.Label(right, text="TEMPO DECORRIDO", font=("Segoe UI", 8),
                 fg=C_MUTED, bg=C_WHITE).pack(anchor=tk.E)
        self.timer_lbl = tk.Label(right, text="00:00:00",
                                  font=("Consolas", 34, "bold"),
                                  fg=C_PRIMARY, bg=C_WHITE)
        self.timer_lbl.pack(anchor=tk.E)
        self.inicio_lbl = tk.Label(right, text="Início:  --:--:--",
                                   font=("Segoe UI", 9), fg=C_MUTED, bg=C_WHITE)
        self.inicio_lbl.pack(anchor=tk.E, pady=(6, 0))
        self.data_lbl = tk.Label(right, text="Data:    --/--/----",
                                 font=("Segoe UI", 9), fg=C_MUTED, bg=C_WHITE)
        self.data_lbl.pack(anchor=tk.E)

    def _build_buttons(self, parent):
        frm = tk.Frame(parent, bg=C_BG)
        frm.pack(fill=tk.X, pady=(0, 10))
        self.btn_iniciar = tk.Button(
            frm, text="▶   INICIAR",
            font=("Segoe UI", 12, "bold"), bg=C_GREEN, fg=C_WHITE,
            activebackground="#1A6B3A", relief=tk.FLAT,
            padx=30, pady=11, cursor="hand2", command=self.iniciar)
        self.btn_iniciar.pack(side=tk.LEFT, padx=(0, 10))
        self.btn_encerrar = tk.Button(
            frm, text="■   ENCERRAR",
            font=("Segoe UI", 12, "bold"), bg=C_GRAY, fg=C_WHITE,
            activebackground=C_RED, relief=tk.FLAT,
            padx=30, pady=11, cursor="hand2",
            command=self.encerrar, state=tk.DISABLED)
        self.btn_encerrar.pack(side=tk.LEFT)

    def _build_obs(self, parent):
        frm = tk.LabelFrame(
            parent, text=" Observações / Tarefa Realizada ",
            font=("Segoe UI", 10, "bold"), fg=C_TEXT,
            bg=C_WHITE, relief=tk.FLAT,
            highlightbackground="#D5D8DC", highlightthickness=1,
            padx=12, pady=8)
        frm.pack(fill=tk.X, pady=(0, 10))
        self.obs_text = tk.Text(
            frm, height=3, font=("Segoe UI", 10), relief=tk.FLAT,
            bg="#FDFEFE", fg=C_GRAY, wrap=tk.WORD,
            highlightbackground="#D5D8DC", highlightthickness=1,
            insertbackground=C_PRIMARY)
        self.obs_text.pack(fill=tk.X)
        self.obs_text.insert("1.0", self._PLACEHOLDER)
        self.obs_text.bind("<FocusIn>",  self._obs_in)
        self.obs_text.bind("<FocusOut>", self._obs_out)

    def _obs_in(self, _):
        if self.obs_text.get("1.0", tk.END).strip() == self._PLACEHOLDER:
            self.obs_text.delete("1.0", tk.END)
            self.obs_text.config(fg=C_TEXT)

    def _obs_out(self, _):
        if not self.obs_text.get("1.0", tk.END).strip():
            self.obs_text.insert("1.0", self._PLACEHOLDER)
            self.obs_text.config(fg=C_GRAY)

    def _build_history(self, parent):
        outer = tk.LabelFrame(
            parent, text=" Histórico de Registros ",
            font=("Segoe UI", 10, "bold"), fg=C_TEXT,
            bg=C_WHITE, relief=tk.FLAT,
            highlightbackground="#D5D8DC", highlightthickness=1,
            padx=10, pady=8)
        outer.pack(fill=tk.BOTH, expand=True)

        # Rodapé — empacotado ANTES da treeview para sempre aparecer
        tk.Label(outer, text=f"Planilha: {EXCEL_FILE}",
                 font=("Segoe UI", 8), fg=C_MUTED, bg=C_WHITE,
                 anchor=tk.W).pack(side=tk.BOTTOM, fill=tk.X)
        tk.Label(outer,
                 text="* Clique com o botão direito sobre o registro para editar ou excluir",
                 font=("Segoe UI", 8), fg=C_MUTED, bg=C_WHITE,
                 anchor=tk.W).pack(side=tk.BOTTOM, fill=tk.X, pady=(4, 0))

        # Treeview — 6 colunas incluindo Data Fim
        cols   = ("D. Início", "Início", "D. Fim", "Fim", "Duração", "Observações")
        widths = (82, 65, 82, 65, 78, 270)
        self.tree = ttk.Treeview(outer, columns=cols, show="headings",
                                 height=6, style="HE.Treeview")
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col)
            anc = tk.W if col == "Observações" else tk.CENTER
            self.tree.column(col, width=w, anchor=anc,
                             stretch=(col == "Observações"))
        self.tree.tag_configure("alt", background=C_ROW_ALT)
        self.tree.bind("<Button-3>", self._on_right_click)

        sb = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

    # ══════════════════════════════════════════════════════════════
    # ABA 2 — RESUMO
    # ══════════════════════════════════════════════════════════════
    def _build_tab_resumo(self, parent):
        f = tk.Frame(parent, bg=C_BG, pady=10)
        f.pack(fill=tk.BOTH, expand=True)

        sel_card = tk.Frame(f, bg=C_WHITE,
                            highlightbackground="#D5D8DC", highlightthickness=1,
                            padx=16, pady=10)
        sel_card.pack(fill=tk.X, pady=(0, 10))
        tk.Label(sel_card, text="Período:", font=("Segoe UI", 10, "bold"),
                 fg=C_TEXT, bg=C_WHITE).pack(side=tk.LEFT, padx=(0, 10))
        self._periods = self._build_period_list(24)
        labels = [p[0] for p in self._periods]
        self.period_var = tk.StringVar(value=labels[0])
        self.period_cb = ttk.Combobox(
            sel_card, textvariable=self.period_var,
            values=labels, state="readonly",
            font=("Segoe UI", 10), width=34)
        self.period_cb.pack(side=tk.LEFT)
        self.period_cb.bind("<<ComboboxSelected>>",
                            lambda _: self._refresh_resumo())

        cards = tk.Frame(f, bg=C_BG)
        cards.pack(fill=tk.X, pady=(0, 10))

        c1 = tk.Frame(cards, bg=C_WHITE,
                      highlightbackground="#D5D8DC", highlightthickness=1,
                      padx=18, pady=14)
        c1.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))
        tk.Label(c1, text="TOTAL DO PERÍODO", font=("Segoe UI", 8, "bold"),
                 fg=C_MUTED, bg=C_WHITE).pack(anchor=tk.W)
        self.lbl_periodo_sub = tk.Label(c1, text="",
                                        font=("Segoe UI", 8), fg=C_MUTED, bg=C_WHITE)
        self.lbl_periodo_sub.pack(anchor=tk.W)
        self.lbl_periodo = tk.Label(c1, text="--:--",
                                    font=("Consolas", 30, "bold"),
                                    fg=C_PRIMARY, bg=C_WHITE)
        self.lbl_periodo.pack(anchor=tk.W, pady=(6, 0))
        tk.Label(c1, text="horas no período  (16 → 15)",
                 font=("Segoe UI", 9), fg=C_MUTED, bg=C_WHITE).pack(anchor=tk.W)

        today = datetime.date.today()
        ws0, we0 = self._semana(today)
        c2 = tk.Frame(cards, bg=C_WHITE,
                      highlightbackground="#D5D8DC", highlightthickness=1,
                      padx=18, pady=14)
        c2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tk.Label(c2, text="SEMANA ATUAL", font=("Segoe UI", 8, "bold"),
                 fg=C_MUTED, bg=C_WHITE).pack(anchor=tk.W)
        tk.Label(c2, text=f"{ws0.strftime('%d/%m')} a {we0.strftime('%d/%m/%Y')}",
                 font=("Segoe UI", 8), fg=C_MUTED, bg=C_WHITE).pack(anchor=tk.W)
        self.lbl_semana = tk.Label(c2, text="--:--",
                                   font=("Consolas", 30, "bold"),
                                   fg=C_BLUE, bg=C_WHITE)
        self.lbl_semana.pack(anchor=tk.W, pady=(6, 0))
        tk.Label(c2, text="horas esta semana",
                 font=("Segoe UI", 9), fg=C_MUTED, bg=C_WHITE).pack(anchor=tk.W)

        det = tk.LabelFrame(
            f, text=" Registros do Período ",
            font=("Segoe UI", 10, "bold"), fg=C_TEXT,
            bg=C_WHITE, relief=tk.FLAT,
            highlightbackground="#D5D8DC", highlightthickness=1,
            padx=10, pady=8)
        det.pack(fill=tk.BOTH, expand=True)
        cols   = ("D. Início", "Início", "D. Fim", "Fim", "Duração", "Observações")
        widths = (82, 65, 82, 65, 78, 265)
        self.tree_res = ttk.Treeview(det, columns=cols, show="headings",
                                     height=8, style="Sum.Treeview")
        for col, w in zip(cols, widths):
            self.tree_res.heading(col, text=col)
            anc = tk.W if col == "Observações" else tk.CENTER
            self.tree_res.column(col, width=w, anchor=anc,
                                 stretch=(col == "Observações"))
        self.tree_res.tag_configure("alt", background=C_ROW_ALT)
        sb2 = ttk.Scrollbar(det, orient=tk.VERTICAL,
                             command=self.tree_res.yview)
        self.tree_res.configure(yscrollcommand=sb2.set)
        self.tree_res.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb2.pack(side=tk.RIGHT, fill=tk.Y)

    # ── Períodos ─────────────────────────────────────────────────
    def _build_period_list(self, n: int = 24):
        today = datetime.date.today()
        if today.day >= 16:
            end_m, end_y = (today.month + 1, today.year) if today.month < 12 \
                           else (1, today.year + 1)
        else:
            end_m, end_y = today.month, today.year
        periods = []
        m, y = end_m, end_y
        for _ in range(n):
            end   = datetime.date(y, m, 15)
            start = datetime.date(y - 1 if m == 1 else y,
                                  12     if m == 1 else m - 1, 16)
            label = f"{start.strftime('%d/%m/%Y')}  -  {end.strftime('%d/%m/%Y')}"
            periods.append((label, start, end))
            m -= 1
            if m == 0:
                m, y = 12, y - 1
        return periods

    def _refresh_resumo(self):
        try:
            idx = [p[0] for p in self._periods].index(self.period_var.get())
        except ValueError:
            idx = 0
        _, ps, pe = self._periods[idx]
        self.lbl_periodo_sub.config(
            text=f"{ps.strftime('%d/%m/%Y')} a {pe.strftime('%d/%m/%Y')}")
        today = datetime.date.today()
        ws0, we0 = self._semana(today)
        per_sec = sem_sec = 0
        period_rows: list = []
        if os.path.exists(EXCEL_FILE):
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE,
                                            read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    d = self._parse_date(row[0])
                    if d is None:
                        continue
                    _, dur_sec_f = self._compute_dur(row[0], row[1], row[2], row[3])
                    sec = int(dur_sec_f * 3600)
                    if ws0 <= d <= we0:
                        sem_sec += sec
                    if ps <= d <= pe:
                        per_sec += sec
                        period_rows.append(row)
                wb.close()
            except Exception:
                pass

        def fmt(sec: int) -> str:
            return f"{sec // 3600:02d}:{(sec % 3600) // 60:02d}"

        self.lbl_periodo.config(text=fmt(per_sec))
        self.lbl_semana.config(text=fmt(sem_sec))
        for item in self.tree_res.get_children():
            self.tree_res.delete(item)
        for i, row in enumerate(sorted(period_rows, key=lambda r: str(r[0]))):
            tag = "alt" if i % 2 == 0 else ""
            # row: [0]d_ini [1]h_ini [2]d_fim [3]h_fim [4]dur [5]dur_dec [6]obs
            self.tree_res.insert("", tk.END,
                                 values=(row[0], row[1], row[2],
                                         row[3], row[4], row[6] or ""),
                                 tags=(tag,))

    def _on_tab_change(self, _):
        if self.nb.index(self.nb.select()) == 1:
            self._refresh_resumo()

    # ── Helpers ───────────────────────────────────────────────────
    @staticmethod
    def _semana(today: datetime.date):
        m = today - datetime.timedelta(days=today.weekday())
        return m, m + datetime.timedelta(days=6)

    @staticmethod
    def _parse_date(val) -> datetime.date | None:
        try:
            return datetime.datetime.strptime(str(val), "%d/%m/%Y").date()
        except Exception:
            return None

    @staticmethod
    def _parse_dur(val) -> int:
        if isinstance(val, (int, float)):          # valor cacheado pelo Excel (fração de dia)
            return int(float(val) * 86400)
        try:
            h, m, s = str(val).split(":")
            return int(h) * 3600 + int(m) * 60 + int(s)
        except Exception:
            return 0

    @staticmethod
    def _compute_dur(d_ini, h_ini, d_fim, h_fim) -> tuple[str, float]:
        """Calcula duração a partir das colunas de data/hora. Retorna (hh:mm:ss, decimal)."""
        try:
            dt_i = datetime.datetime.strptime(f"{d_ini} {h_ini}", "%d/%m/%Y %H:%M:%S")
            dt_f = datetime.datetime.strptime(f"{d_fim} {h_fim}", "%d/%m/%Y %H:%M:%S")
            total = int((dt_f - dt_i).total_seconds())
            h, r  = divmod(total, 3600)
            m, s  = divmod(r, 60)
            return f"{h:02d}:{m:02d}:{s:02d}", round(total / 3600, 4)
        except Exception:
            return "00:00:00", 0.0

    @staticmethod
    def _dur_formulas(r: int) -> tuple[str, str]:
        """Retorna as fórmulas Excel de duração para a linha r."""
        inner = (f"(DATE(RIGHT(C{r},4),MID(C{r},4,2),LEFT(C{r},2))+TIMEVALUE(D{r}))"
                 f"-(DATE(RIGHT(A{r},4),MID(A{r},4,2),LEFT(A{r},2))+TIMEVALUE(B{r}))")
        return f"={inner}", f"=({inner})*24"

    # ══════════════════════════════════════════════════════════════
    # TIMER
    # ══════════════════════════════════════════════════════════════
    def iniciar(self):
        self.inicio = datetime.datetime.now()
        self.elapsed_seconds = 0
        self.timer_running = True
        self.status_dot.config(fg=C_GREEN)
        self.status_lbl.config(text="Em andando...", fg=C_GREEN)
        self.inicio_lbl.config(text=f"Início:  {self.inicio.strftime('%H:%M:%S')}")
        self.data_lbl.config(text=f"Data:    {self.inicio.strftime('%d/%m/%Y')}")
        self.btn_iniciar.config(state=tk.DISABLED, bg=C_GRAY)
        self.btn_encerrar.config(state=tk.NORMAL, bg=C_RED)
        self.obs_text.config(fg=C_TEXT)
        self.obs_text.delete("1.0", tk.END)
        self._tick()

    def _tick(self):
        if not self.timer_running:
            return
        self.elapsed_seconds += 1
        h, r = divmod(self.elapsed_seconds, 3600)
        m, s = divmod(r, 60)
        self.timer_lbl.config(text=f"{h:02d}:{m:02d}:{s:02d}")
        self.root.after(1000, self._tick)

    def encerrar(self):
        if not self.inicio:
            return
        fim = datetime.datetime.now()
        self.timer_running = False
        raw = self.obs_text.get("1.0", tk.END).strip()
        obs = "" if raw == self._PLACEHOLDER else raw
        total = int((fim - self.inicio).total_seconds())
        h, r = divmod(total, 3600)
        m, s = divmod(r, 60)
        dur_str = f"{h:02d}:{m:02d}:{s:02d}"
        dur_dec = round(total / 3600, 4)
        self._save_excel(self.inicio, fim, dur_str, dur_dec, obs)
        self._tree_insert(
            self.inicio.strftime("%d/%m/%Y"),
            self.inicio.strftime("%H:%M:%S"),
            fim.strftime("%d/%m/%Y"),
            fim.strftime("%H:%M:%S"),
            dur_str, obs)
        messagebox.showinfo("Registro Salvo",
                            f"Horas extras registradas!\n\n"
                            f"Duração: {dur_str}  ({dur_dec} h)\n\n"
                            f"Arquivo:\n{EXCEL_FILE}")
        self._reset_ui()

    def _reset_ui(self):
        self.inicio = None
        self.elapsed_seconds = 0
        self.status_dot.config(fg=C_GRAY)
        self.status_lbl.config(text="Aguardando início", fg=C_MUTED)
        self.inicio_lbl.config(text="Início:  --:--:--")
        self.data_lbl.config(text="Data:    --/--/----")
        self.timer_lbl.config(text="00:00:00")
        self.btn_iniciar.config(state=tk.NORMAL, bg=C_GREEN)
        self.btn_encerrar.config(state=tk.DISABLED, bg=C_GRAY)
        self.obs_text.delete("1.0", tk.END)
        self.obs_text.insert("1.0", self._PLACEHOLDER)
        self.obs_text.config(fg=C_GRAY)

    # ══════════════════════════════════════════════════════════════
    # EDITAR REGISTRO
    # ══════════════════════════════════════════════════════════════
    def _editar_registro(self):
        sel = self.tree.selection()
        if not sel:
            return
        item = sel[0]
        vals = self.tree.item(item, "values")
        # vals = (d_ini, h_ini, d_fim, h_fim, dur, obs)

        dlg = tk.Toplevel(self.root)
        dlg.title("Editar Registro")
        dlg.geometry("500x370")
        dlg.resizable(False, False)
        dlg.configure(bg=C_BG)
        dlg.grab_set()
        dlg.transient(self.root)
        dlg.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width()  - 500) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 370) // 2
        dlg.geometry(f"500x370+{x}+{y}")

        tk.Frame(dlg, bg=C_PRIMARY, height=4).pack(fill=tk.X)
        tk.Label(dlg, text="Editar Registro",
                 font=("Segoe UI", 13, "bold"), fg=C_TEXT, bg=C_BG
                 ).pack(pady=(14, 10))

        form = tk.Frame(dlg, bg=C_BG, padx=28)
        form.pack(fill=tk.X)
        form.columnconfigure(1, weight=1)

        def lbl(text, r):
            tk.Label(form, text=text, font=("Segoe UI", 9, "bold"),
                     fg=C_TEXT, bg=C_BG, anchor=tk.W
                     ).grid(row=r, column=0, sticky=tk.W, pady=5, padx=(0, 10))

        def ent(r, val, w=18):
            e = tk.Entry(form, font=("Segoe UI", 10), relief=tk.FLAT,
                         highlightbackground="#D5D8DC", highlightthickness=1,
                         bg=C_WHITE, width=w)
            e.insert(0, val)
            e.grid(row=r, column=1, sticky=tk.W, pady=5)
            return e

        def hint(text, r):
            tk.Label(form, text=text, font=("Segoe UI", 8),
                     fg=C_MUTED, bg=C_BG).grid(row=r, column=2,
                                                padx=(8, 0), sticky=tk.W)

        lbl("Data início:",  0); e_d_ini = ent(0, vals[0]); hint("dd/mm/aaaa", 0)
        lbl("Hora início:",  1); e_h_ini = ent(1, vals[1]); hint("hh:mm:ss",   1)
        lbl("Data fim:",     2); e_d_fim = ent(2, vals[2]); hint("dd/mm/aaaa", 2)
        lbl("Hora fim:",     3); e_h_fim = ent(3, vals[3]); hint("hh:mm:ss",   3)
        lbl("Observações:",  4)
        e_obs = tk.Text(form, height=3, font=("Segoe UI", 10), relief=tk.FLAT,
                        highlightbackground="#D5D8DC", highlightthickness=1,
                        bg=C_WHITE, width=30)
        e_obs.insert("1.0", vals[5])
        e_obs.grid(row=4, column=1, columnspan=2, sticky=tk.EW, pady=5)

        def salvar():
            d_ini_v = e_d_ini.get().strip()
            h_ini_v = e_h_ini.get().strip()
            d_fim_v = e_d_fim.get().strip()
            h_fim_v = e_h_fim.get().strip()
            ob_v    = e_obs.get("1.0", tk.END).strip()
            try:
                dt_i = datetime.datetime.strptime(
                    f"{d_ini_v} {h_ini_v}", "%d/%m/%Y %H:%M:%S")
                dt_f = datetime.datetime.strptime(
                    f"{d_fim_v} {h_fim_v}", "%d/%m/%Y %H:%M:%S")
                if dt_f <= dt_i:
                    messagebox.showerror(
                        "Data inválida",
                        "A data/hora de fim deve ser posterior ao início.",
                        parent=dlg)
                    return
                total = int((dt_f - dt_i).total_seconds())
                hh, rr = divmod(total, 3600)
                mm, ss = divmod(rr, 60)
                dur_str = f"{hh:02d}:{mm:02d}:{ss:02d}"
                dur_dec = round(total / 3600, 4)
            except ValueError:
                messagebox.showerror(
                    "Formato inválido",
                    "Data: dd/mm/aaaa\nHora: hh:mm:ss",
                    parent=dlg)
                return
            # Valores antigos para localizar no Excel
            old_d_ini, old_h_ini, old_h_fim = vals[0], vals[1], vals[3]
            self._excel_update_row(old_d_ini, old_h_ini, old_h_fim,
                                   d_ini_v, h_ini_v, d_fim_v, h_fim_v,
                                   dur_str, dur_dec, ob_v)
            self.tree.item(item, values=(d_ini_v, h_ini_v,
                                         d_fim_v, h_fim_v,
                                         dur_str, ob_v))
            dlg.destroy()

        bf = tk.Frame(dlg, bg=C_BG, pady=14)
        bf.pack()
        tk.Button(bf, text="  Salvar  ", font=("Segoe UI", 10, "bold"),
                  bg=C_GREEN, fg=C_WHITE, relief=tk.FLAT,
                  padx=20, pady=7, cursor="hand2",
                  command=salvar).pack(side=tk.LEFT, padx=8)
        tk.Button(bf, text="Cancelar", font=("Segoe UI", 10),
                  bg=C_GRAY, fg=C_WHITE, relief=tk.FLAT,
                  padx=20, pady=7, cursor="hand2",
                  command=dlg.destroy).pack(side=tk.LEFT, padx=8)

    def _excel_update_row(self, old_d_ini, old_h_ini, old_h_fim,
                          new_d_ini, new_h_ini, new_d_fim, new_h_fim,
                          dur_str, dur_dec, obs):
        if not os.path.exists(EXCEL_FILE):
            return
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                # col[0]=d_ini, col[1]=h_ini, col[3]=h_fim
                if (str(row[0].value) == str(old_d_ini) and
                        str(row[1].value) == str(old_h_ini) and
                        str(row[3].value) == str(old_h_fim)):
                    row[0].value = new_d_ini
                    row[1].value = new_h_ini
                    row[2].value = new_d_fim
                    row[3].value = new_h_fim
                    row[4].value = dur_str
                    row[5].value = dur_dec
                    row[6].value = obs
                    break
            wb.save(EXCEL_FILE)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível atualizar a planilha:\n{e}")

    # ══════════════════════════════════════════════════════════════
    # EXCLUIR REGISTRO
    # ══════════════════════════════════════════════════════════════
    def excluir_registro(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Nenhum registro selecionado",
                                   "Clique com o botão direito em um registro.")
            return
        vals = self.tree.item(sel[0], "values")
        # vals = (d_ini, h_ini, d_fim, h_fim, dur, obs)
        ok = messagebox.askyesno(
            "Confirmar Exclusão",
            f"Excluir este registro?\n\n"
            f"  Início:  {vals[0]} {vals[1]}\n"
            f"  Fim:     {vals[2]} {vals[3]}\n"
            f"  Duração: {vals[4]}\n\n"
            "Esta ação não pode ser desfeita.")
        if not ok:
            return
        self.tree.delete(sel[0])
        self._excel_delete_row(vals[0], vals[1], vals[3])
        self._recolor(self.tree)

    def _excel_delete_row(self, d_ini, h_ini, h_fim):
        if not os.path.exists(EXCEL_FILE):
            return
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                if (str(row[0].value) == str(d_ini) and
                        str(row[1].value) == str(h_ini) and
                        str(row[3].value) == str(h_fim)):
                    ws.delete_rows(row[0].row)
                    break
            wb.save(EXCEL_FILE)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível atualizar a planilha:\n{e}")

    def _recolor(self, tree):
        for i, item in enumerate(tree.get_children()):
            tree.item(item, tags=("alt" if i % 2 == 0 else "",))

    # ══════════════════════════════════════════════════════════════
    # EXCEL
    # ══════════════════════════════════════════════════════════════
    def _excel_headers(self, ws):
        hdrs   = ["Data Início", "Hora Início", "Data Fim", "Hora Fim",
                  "Duração (hh:mm:ss)", "Duração (h decimal)", "Observações"]
        widths = [14, 14, 14, 14, 22, 22, 55]
        for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = PatternFill("solid", fgColor="1A252F")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"

    def _excel_style_row(self, ws, row_idx: int, n_cols: int):
        for ci in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=ci)
            last_col = n_cols
            cell.alignment = Alignment(
                horizontal="center" if ci < last_col else "left",
                vertical="center",
                wrap_text=(ci == last_col))
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EBF5FB")
            cell.border = Border(bottom=Side(style="thin", color="D5D8DC"))
        ws.row_dimensions[row_idx].height = 18

    def _save_excel(self, inicio: datetime.datetime,
                    fim: datetime.datetime,
                    dur_str: str, dur_dec: float, obs: str):
        if os.path.exists(EXCEL_FILE):
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Horas Extras"
            self._excel_headers(ws)

        row_idx = ws.max_row + 1
        # Colunas A-D e G: valores diretos
        for ci, v in enumerate([
            inicio.strftime("%d/%m/%Y"),
            inicio.strftime("%H:%M:%S"),
            fim.strftime("%d/%m/%Y"),
            fim.strftime("%H:%M:%S"),
        ], 1):
            ws.cell(row=row_idx, column=ci, value=v)
        ws.cell(row=row_idx, column=7, value=obs)

        # Colunas E e F: fórmulas automáticas
        fe, ff = self._dur_formulas(row_idx)
        ce = ws.cell(row=row_idx, column=5, value=fe)
        ce.number_format = '[h]:mm:ss'
        cf = ws.cell(row=row_idx, column=6, value=ff)
        cf.number_format = '0.0000'

        self._excel_style_row(ws, row_idx, 7)
        wb.save(EXCEL_FILE)

    # ── Migração: formato antigo (6 cols) → novo (7 cols) ────────
    def _migrate_excel_if_needed(self):
        if not os.path.exists(EXCEL_FILE):
            return
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            header_a1 = str(ws.cell(1, 1).value or "")
            # Novo formato já tem "Data Início" na célula A1
            if header_a1 == "Data Início":
                wb.close()
                return
            # Lê todas as linhas de dados
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    rows.append(row)
            # Reconstrói a planilha com novo formato
            ws.delete_rows(1, ws.max_row + 1)
            self._excel_headers(ws)
            for old_row in rows:
                # old: [0]d_ini [1]h_ini [2]h_fim [3]dur_str [4]dur_dec [5]obs
                d_ini, h_ini, h_fim = str(old_row[0]), str(old_row[1]), str(old_row[2])
                dur_str = str(old_row[3])
                dur_dec = old_row[4]
                obs     = old_row[5] or ""
                # Calcula data fim a partir do início + duração
                try:
                    dt_i  = datetime.datetime.strptime(f"{d_ini} {h_ini}",
                                                       "%d/%m/%Y %H:%M:%S")
                    h, m, s = dur_str.split(":")
                    delta = datetime.timedelta(seconds=int(h)*3600 +
                                               int(m)*60 + int(s))
                    d_fim = (dt_i + delta).strftime("%d/%m/%Y")
                except Exception:
                    d_fim = d_ini
                row_idx = ws.max_row + 1
                new_vals = [d_ini, h_ini, d_fim, h_fim, dur_str, dur_dec, obs]
                for ci, v in enumerate(new_vals, 1):
                    ws.cell(row=row_idx, column=ci, value=v)
                self._excel_style_row(ws, row_idx, 7)
            wb.save(EXCEL_FILE)
        except Exception:
            pass

    # ── Treeview helpers ─────────────────────────────────────────
    def _tree_insert(self, d_ini, h_ini, d_fim, h_fim, dur, obs):
        tag = "alt" if len(self.tree.get_children()) % 2 == 0 else ""
        self.tree.insert("", 0,
                         values=(d_ini, h_ini, d_fim, h_fim, dur, obs),
                         tags=(tag,))

    def _load_history(self):
        if not os.path.exists(EXCEL_FILE):
            return
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE,
                                        read_only=True, data_only=True)
            ws = wb.active
            rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                    if r[0]]
            for i, row in enumerate(reversed(rows)):
                tag = "alt" if i % 2 == 0 else ""
                # Duração sempre calculada das colunas de data/hora (ignora E/F)
                dur_str, _ = self._compute_dur(row[0], row[1], row[2], row[3])
                self.tree.insert("", tk.END,
                                 values=(row[0], row[1], row[2],
                                         row[3], dur_str, row[6] or ""),
                                 tags=(tag,))
            wb.close()
        except Exception:
            pass


if __name__ == "__main__":
    root = tk.Tk()
    ControleHE(root)
    root.mainloop()
